from __future__ import annotations

from math import sqrt
from pathlib import Path
from statistics import mean
import openpyxl


def norm(name: str) -> str:
    return "".join(ch for ch in str(name).lower() if ch.isalnum())


def variance(values: list[float], avg: float) -> float:
    if len(values) < 2:
        return 0.0
    return sum((v - avg) ** 2 for v in values) / (len(values) - 1)


def to_float(value, default=0.0) -> float:
    if value in (None, "", "-"):
        return default
    if isinstance(value, bool):
        return 1.0 if value else 0.0
    s = str(value).strip().lower()
    if s in {"yes", "y", "true", "t", "1", "seizure", "sz", "✓"}:
        return 1.0
    if s in {"no", "n", "false", "f", "0", "ok", "none", "—", "-", "✗"}:
        return 0.0
    try:
        return float(s)
    except Exception:
        return default


def analyze_factor(name: str, seizure_values: list[float], normal_values: list[float], threshold: float = 0.2):
    seizure_avg = mean(seizure_values) if seizure_values else 0.0
    normal_avg = mean(normal_values) if normal_values else 0.0
    diff = abs(seizure_avg - normal_avg)

    if len(seizure_values) < 30 or len(normal_values) < 30:
        is_trigger = diff >= threshold
        weight = diff
        method = "threshold"
    else:
        s_var = variance(seizure_values, seizure_avg)
        n_var = variance(normal_values, normal_avg)
        sp = (s_var / len(seizure_values)) + (n_var / len(normal_values))
        if sp == 0:
            t_stat = 0.0
        else:
            t_stat = diff / sqrt(sp)
        is_trigger = t_stat > 2.0
        s_sd = sqrt(s_var)
        weight = (diff / (s_sd + 0.001)) if is_trigger else 0.0
        method = f"welch_t(t={t_stat:.2f})"

    return {
        "factor": name,
        "is_trigger": is_trigger,
        "seizure_avg": seizure_avg,
        "normal_avg": normal_avg,
        "diff": diff,
        "weight": weight,
        "method": method,
    }


def analyze_sheet(ws):
    headers = [c.value for c in ws[1]]
    header_map = {norm(h): i for i, h in enumerate(headers) if h is not None}

    def idx(*names: str):
        for n in names:
            if norm(n) in header_map:
                return header_map[norm(n)]
        return None

    i_date = idx("date")
    i_status = idx("status", "hadseizure")
    i_sleep_h = idx("sleeph", "sleephours")
    i_sleep_q = idx("sleepq", "sleepquality")
    i_sleep_int = idx("sint", "sleepinterruptions")
    i_stress = idx("stress", "stresslevel")
    i_diet = idx("diet", "dietquality")
    i_meds = idx("meds", "medication", "medicationadherence", "medicationtaken")
    i_drug = idx("drug", "druguse")
    i_horm = idx("horm", "hormonal", "hormonalchanges")
    i_sz = idx("sz", "sznum", "seizurecount")

    required = {
        "date": i_date,
        "sleep_h": i_sleep_h,
        "sleep_q": i_sleep_q,
        "sleep_int": i_sleep_int,
        "stress": i_stress,
        "diet": i_diet,
        "meds": i_meds,
        "drug": i_drug,
        "horm": i_horm,
    }
    missing = [k for k, v in required.items() if v is None]
    if missing:
        return {
            "ok": False,
            "missing": missing,
            "headers": headers,
        }

    rows = []
    for r in ws.iter_rows(min_row=3, values_only=True):
        if all(v in (None, "") for v in r):
            continue

        status_txt = str(r[i_status]).strip().lower() if i_status is not None and r[i_status] is not None else ""
        status_bool = to_float(r[i_status], 0.0) if i_status is not None else 0.0
        sz_num = to_float(r[i_sz], 0.0) if i_sz is not None else 0.0
        is_seizure = ("seiz" in status_txt) or (status_bool > 0) or (sz_num > 0)

        rows.append(
            {
                "date": str(r[i_date]),
                "is_seizure": is_seizure,
                "sleep_h": to_float(r[i_sleep_h]),
                "sleep_q": to_float(r[i_sleep_q]),
                "sleep_int": to_float(r[i_sleep_int]),
                "stress": to_float(r[i_stress]),
                "diet": to_float(r[i_diet]),
                "meds": to_float(r[i_meds]),
                "drug": to_float(r[i_drug]),
                "horm": to_float(r[i_horm]),
            }
        )

    seizure_rows = [x for x in rows if x["is_seizure"]]
    normal_rows = [x for x in rows if not x["is_seizure"]]

    factors = [
        ("Sleep Hours", "sleep_h", 0.2),
        ("Sleep Quality", "sleep_q", 0.2),
        ("Sleep Interruptions", "sleep_int", 0.2),
        ("Stress Level", "stress", 0.2),
        ("Diet Quality", "diet", 0.2),
        ("Medication", "meds", 0.2),
        ("Drug Use", "drug", 0.2),
        ("Hormonal Changes", "horm", 0.2),
    ]

    results = []
    for label, key, threshold in factors:
        seizure_vals = [x[key] for x in seizure_rows]
        normal_vals = [x[key] for x in normal_rows]
        results.append(analyze_factor(label, seizure_vals, normal_vals, threshold))

    results.sort(key=lambda x: x["weight"], reverse=True)

    return {
        "ok": True,
        "rows": len(rows),
        "seizure_rows": len(seizure_rows),
        "normal_rows": len(normal_rows),
        "results": results,
    }


def main() -> None:
    xlsx_path = Path("seizure_patient_data.xlsx")
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    print("=== Workbook Test Run ===")
    print(f"Profiles found: {len(wb.sheetnames) - (1 if 'INDEX' in wb.sheetnames else 0)}")

    for sheet_name in wb.sheetnames:
        if sheet_name.upper() == "INDEX":
            continue

        ws = wb[sheet_name]
        output = analyze_sheet(ws)

        print(f"\n--- {sheet_name} ---")
        if not output["ok"]:
            print(f"Missing required columns: {output['missing']}")
            continue

        print(
            f"Rows={output['rows']} | Seizure={output['seizure_rows']} | Normal={output['normal_rows']}"
        )

        top = output["results"][:4]
        for r in top:
            tag = "TRIGGER" if r["is_trigger"] else "not-trigger"
            print(
                f"{r['factor']:<20} {tag:<12} "
                f"diff={r['diff']:.3f} weight={r['weight']:.3f} via {r['method']}"
            )


if __name__ == "__main__":
    main()
